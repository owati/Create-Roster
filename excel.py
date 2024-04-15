from openpyxl import Workbook
import json

data  = {}
with open("data.json") as f:
    data = json.load(f)

wb = Workbook()
ws = wb.active

ws['A1'] = "Date"
ws['B1'] = "Stage"
ws['C1'] = "Recording"
ws['D1'] = "Mixer"
ws['E1'] = "Service"

count = 1
for date in data:
    stage = data[date].get("stage")
    if isinstance(stage, list):
        stage = ", ".join(stage)

    mixer = data[date].get("mixer")
    if isinstance(mixer, list):
        mixer = ", ".join(mixer)

    service = ""
    match (count % 3):
        case 1:
            service = "Monday: Bible Study"
        case 2:
            service = "Thursday: Prayer Meeting"
        case 0:
            service = "Sunday: Service"

    ws.append([date, stage, data[date].get("recording"), mixer, service])

    count += 1

wb.save("Organising Allocation.xlsx")